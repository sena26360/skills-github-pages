import openpyxl
from datetime import datetime
from flask import Flask, request, render_template

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        files = request.files.getlist('files')  # 여러 파일 가져오기
        cell = request.form['cell']
        new_value = request.form['new_value']
        data_type = request.form['data_type']

        for file in files:
            # 엑셀 파일 불러오기 (기존 서식 유지)
            wb = openpyxl.load_workbook(file)
            ws = wb.active  # 기본 시트 선택

            # 셀 값 변경
            if data_type == 'text':
                ws[cell] = new_value  # 문자열 변경
            elif data_type == 'date':
                try:
                    ws[cell] = datetime.strptime(new_value, "%Y-%m-%d")  # 날짜 변경
                except ValueError:
                    return "날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식으로 입력하세요."

            # 수정된 엑셀 파일 저장
            modified_file_name = f'0.{file.filename}'
            wb.save(modified_file_name)

        return '모든 파일이 성공적으로 수정되었습니다.'

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
